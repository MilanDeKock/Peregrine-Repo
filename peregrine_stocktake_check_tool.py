import io
import os
import re
import tempfile
import time
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import pdfplumber
import camelot
import requests
import streamlit as st
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from pytz import timezone

# -------------------------
# Configuration
# -------------------------
load_dotenv()
TZ = timezone("Africa/Johannesburg")

API_BASE_URL = "https://inventory.dearsystems.com/ExternalApi/v2"
CALLS_PER_MINUTE = 50
CALL_INTERVAL = 60.0 / CALLS_PER_MINUTE

st.set_page_config(
    page_title="Peregrine Stocktake Check Tool",
    page_icon="\U0001F4E6",
    layout="wide",
)

st.title("\U0001F4E6 Peregrine Stocktake Check Tool")
st.caption("Upload a Stocktake PDF \u2192 Inventory pulled from Cin7 Core \u2192 Variance Analysis XLSX download")


# -------------------------
# Cin7 Core API Client
# -------------------------
class Cin7Client:
    """Cin7 Core API client with built-in rate limiting."""

    def __init__(self, account_id: str, api_key: str):
        self.headers = {
            "Content-Type": "application/json",
            "api-auth-accountid": account_id,
            "api-auth-applicationkey": api_key,
        }
        self._last_call_time = 0.0

    def _throttle(self):
        elapsed = time.time() - self._last_call_time
        if elapsed < CALL_INTERVAL:
            time.sleep(CALL_INTERVAL - elapsed)
        self._last_call_time = time.time()

    def get(self, endpoint, params=None):
        self._throttle()
        url = f"{API_BASE_URL}/{endpoint}"
        for attempt in range(3):
            resp = requests.get(url, headers=self.headers, params=params, timeout=30)
            if resp.status_code in (429, 503):
                # Rate limited — back off and retry
                time.sleep(60)
                self._last_call_time = time.time()
                continue
            break
        resp.raise_for_status()
        return resp.json()


def get_api_credentials():
    """Load API credentials from st.secrets (Streamlit Cloud) or .env (local)."""
    account_id = None
    api_key = None
    try:
        account_id = st.secrets["CIN7_ACCOUNT_ID"]
        api_key = st.secrets["CIN7_API_KEY"]
    except (KeyError, FileNotFoundError, AttributeError):
        pass
    if not account_id or not api_key:
        account_id = os.getenv("CIN7_ACCOUNT_ID")
        api_key = os.getenv("CIN7_API_KEY")
    return account_id, api_key


# -------------------------
# API Data Fetchers
# -------------------------
@st.cache_data(ttl=300, show_spinner=False)
def fetch_inventory_from_api(account_id, api_key):
    """Fetch all products from Cin7 Core API with pagination."""
    client = Cin7Client(account_id, api_key)
    all_products = []
    page = 1
    limit = 1000

    while True:
        data = client.get("product", params={"Page": page, "Limit": limit})
        products = data.get("Products", [])
        all_products.extend(products)
        total = data.get("Total", 0)
        if page * limit >= total:
            break
        page += 1

    rows = []
    for p in all_products:
        rows.append({
            "ProductCode": str(p.get("SKU", "")).strip(),
            "AverageCost": p.get("AverageCost", 0) or 0,
            "Category": p.get("Category", "") or "",
            "DefaultUnitOfMeasure": p.get("UOM", "") or "",
            "CommaDelimitedTags": p.get("Tags", "") or "",
        })
    return pd.DataFrame(rows)


# -------------------------
# Bulk Reference Data Fetchers
# -------------------------
def fetch_all_soh(client, progress_cb=None):
    """
    Fetch SOH for ALL products by paginating /ref/productavailability
    without a SKU filter. Returns dict keyed by SKU (upper).
    """
    soh_data = {}
    page = 1
    limit = 1000
    total_fetched = 0

    while True:
        try:
            data = client.get("ref/productavailability", params={
                "Page": page, "Limit": limit,
            })
        except requests.HTTPError:
            break
        rows = data.get("ProductAvailabilityList", [])
        if not rows:
            break

        # Group rows by SKU
        for row in rows:
            sku_key = (row.get("SKU") or "").strip().upper()
            if not sku_key:
                continue
            if sku_key not in soh_data:
                soh_data[sku_key] = {
                    "SOH_Total_OnHand": 0,
                    "SOH_Total_Available": 0,
                    "locations": [],
                }
            oh = row.get("OnHand", 0) or 0
            av = row.get("Available", 0) or 0
            loc = row.get("Location", "?")
            soh_data[sku_key]["SOH_Total_OnHand"] += oh
            soh_data[sku_key]["SOH_Total_Available"] += av
            soh_data[sku_key]["locations"].append(f"{loc}: {oh}")

        total_fetched += len(rows)
        total = data.get("Total", 0)
        if progress_cb:
            progress_cb(total_fetched, total)
        if page * limit >= total:
            break
        page += 1

    # Flatten location lists into strings
    for sku_key, entry in soh_data.items():
        entry["SOH_Per_Location"] = " | ".join(entry.pop("locations"))

    return soh_data


def fetch_bom_reverse_index(client, progress_cb=None):
    """
    Fetch ALL products with BOM data and build a reverse index:
    component SKU -> list of parent products that use it.
    """
    bom_index = defaultdict(list)
    page = 1
    limit = 1000

    while True:
        try:
            data = client.get("product", params={
                "Page": page, "Limit": limit, "IncludeBOM": "true",
            })
        except requests.HTTPError:
            break
        products = data.get("Products", [])
        for p in products:
            if p.get("BillOfMaterial") and p.get("BillOfMaterialsProducts"):
                parent_sku = p.get("SKU", "")
                parent_name = p.get("Name", "")
                for comp in p.get("BillOfMaterialsProducts", []):
                    comp_sku = comp.get("ProductCode", "").strip().upper()
                    if comp_sku:
                        bom_index[comp_sku].append({
                            "ParentSKU": parent_sku,
                            "ParentName": parent_name,
                            "Quantity": comp.get("Quantity", 0),
                            "QuantityToProduce": p.get("QuantityToProduce", 1),
                        })
        total = data.get("Total", 0)
        if progress_cb:
            progress_cb(page * limit, total)
        if page * limit >= total:
            break
        page += 1

    return bom_index


def fetch_all_stock_take_lines(client, progress_cb=None, num_takes=5):
    """
    Fetch last N completed stock takes and store ALL line items
    in a dict keyed by SKU (upper). Each SKU gets data from its
    most recent stock take only.
    """
    # Fetch stock take list
    all_takes = []
    page = 1
    page_size = 100

    while True:
        try:
            data = client.get("stockTakeList", params={
                "Status": "COMPLETED", "Limit": str(page_size), "Page": str(page),
            })
        except requests.HTTPError:
            break
        takes = data.get("StockAdjustmentList", [])
        all_takes.extend(takes)
        total = data.get("Total", 0)
        if page * page_size >= total:
            break
        page += 1

    def parse_date(st_entry):
        raw = st_entry.get("EffectiveDate", "") or ""
        try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00"))
        except (ValueError, TypeError):
            return datetime.min

    all_takes.sort(key=parse_date, reverse=True)
    recent_takes = all_takes[:num_takes]

    # Fetch detail for each and aggregate ALL lines
    last_count = {}
    for i, st_entry in enumerate(recent_takes):
        task_id = st_entry.get("TaskID")
        st_number = st_entry.get("StocktakeNumber", "?")
        st_date = st_entry.get("EffectiveDate", "")[:10]
        st_location = st_entry.get("Location", "")

        try:
            detail = client.get("stocktake", params={"TaskID": task_id})
        except requests.HTTPError:
            detail = None

        if progress_cb:
            progress_cb(i + 1, len(recent_takes))

        if not detail:
            continue

        sku_agg = defaultdict(lambda: {"system_soh": 0, "counted": 0})
        for line in detail.get("NonZeroStockOnHandProducts", []):
            sku_key = line.get("SKU", "").strip().upper()
            if sku_key:
                sku_agg[sku_key]["system_soh"] += (line.get("QuantityOnHand", 0) or 0)
                sku_agg[sku_key]["counted"] += (line.get("Adjustment", 0) or 0)

        for line in detail.get("ZeroStockOnHandProducts", []):
            sku_key = line.get("SKU", "").strip().upper()
            if sku_key:
                sku_agg[sku_key]["counted"] += (line.get("Quantity", 0) or 0)

        for sku_key, agg in sku_agg.items():
            if sku_key not in last_count:
                system_soh = agg["system_soh"]
                counted_qty = agg["counted"]
                variance = counted_qty - system_soh
                last_count[sku_key] = {
                    "Last_System_SOH": system_soh,
                    "Last_Counted_Qty": counted_qty,
                    "Last_ST_Variance": variance,
                    "Last_StockTake_Ref": st_number,
                    "Last_StockTake_Date": st_date,
                    "Last_StockTake_Location": st_location,
                }

    return last_count


def fetch_po_data_via_movements(client, skus, progress_cb=None):
    """
    Fetch latest PO receiving for given SKUs using the Product Movements API.
    One call per SKU: GET /product?Sku={SKU}&IncludeMovements=true
    Filters movements for Type 'Purchase' or 'Advanced Purchase',
    picks the most recent one.
    Returns (receiving_data, debug_info).
    """
    receiving_data = {}
    debug = {
        "skus_queried": 0, "skus_with_po": 0,
        "total_movements_seen": 0,
        "errors": [],
    }

    purchase_types = {"purchase", "advanced purchase"}

    for i, sku in enumerate(skus):
        try:
            data = client.get("product", params={
                "Sku": sku, "IncludeMovements": "true",
            })
        except requests.HTTPError as e:
            debug["errors"].append(f"product?Sku={sku}: {e}")
            debug["skus_queried"] += 1
            if progress_cb:
                progress_cb(i + 1, len(skus))
            continue

        products = data.get("Products", [])
        if not products:
            debug["skus_queried"] += 1
            if progress_cb:
                progress_cb(i + 1, len(skus))
            continue

        movements = products[0].get("Movements", [])
        debug["total_movements_seen"] += len(movements)

        # Filter for purchase movements only
        po_movements = [
            m for m in movements
            if (m.get("Type") or "").lower() in purchase_types
        ]

        if po_movements:
            # Sort by date descending to get latest
            def parse_date(m):
                raw = (m.get("Date") or "")[:19]
                try:
                    return datetime.fromisoformat(raw)
                except (ValueError, TypeError):
                    return datetime.min

            po_movements.sort(key=parse_date, reverse=True)
            latest = po_movements[0]
            receiving_data[sku.upper()] = {
                "Last_PO_Number": latest.get("Number", ""),
                "Last_PO_Date": (latest.get("Date") or "")[:10],
                "Last_PO_Qty": latest.get("Quantity", 0) or 0,
                "Last_PO_Location": latest.get("Location") or "",
            }
            debug["skus_with_po"] += 1

        debug["skus_queried"] += 1
        if progress_cb:
            progress_cb(i + 1, len(skus))

    return receiving_data, debug


# -------------------------
# Reference Data Loader
# -------------------------
def load_reference_data(client, progress_bar, status_text):
    """
    Orchestrate bulk fetch of reference data (SOH, BOM, stock takes).
    Returns dict with all pre-loaded data.
    """
    # Phase 1: SOH (0% - 35%)
    status_text.text("Loading SOH for all products...")
    def soh_progress(fetched, total):
        pct = min(0.35 * (fetched / max(total, 1)), 0.34)
        progress_bar.progress(pct, text=f"SOH: {fetched:,} / {total:,} rows")

    soh_data = fetch_all_soh(client, progress_cb=soh_progress)
    progress_bar.progress(0.35, text=f"SOH loaded: {len(soh_data):,} SKUs")

    # Phase 2: BOM (35% - 65%)
    status_text.text("Building BOM reverse index...")
    def bom_progress(fetched, total):
        pct = 0.35 + min(0.30 * (fetched / max(total, 1)), 0.29)
        progress_bar.progress(pct, text="BOM: fetching products...")

    bom_index = fetch_bom_reverse_index(client, progress_cb=bom_progress)
    progress_bar.progress(0.65, text=f"BOM loaded: {len(bom_index):,} components")

    # Phase 3: Stock takes (65% - 100%)
    status_text.text("Fetching stock take history...")
    def st_progress(done, total):
        pct = 0.65 + min(0.35 * (done / max(total, 1)), 0.34)
        progress_bar.progress(pct, text=f"Stock takes: {done}/{total} details")

    stock_take_data = fetch_all_stock_take_lines(client, progress_cb=st_progress)
    progress_bar.progress(1.0, text="Reference data loaded.")

    return {
        "soh_data": soh_data,
        "bom_index": bom_index,
        "stock_take_data": stock_take_data,
        "loaded_at": datetime.now(TZ).strftime("%H:%M:%S"),
    }


# -------------------------
# Investigation Builder (no API calls)
# -------------------------
def build_investigation_from_ref(ref_data, skus, sku_descriptions=None,
                                 sku_costs=None, po_data=None):
    """
    Build investigation + BOM DataFrames from pre-loaded reference data.
    Pure filtering — no API calls.
    PO data is optional (fetched separately on-demand).
    Returns (df_investigation, df_bom).
    """
    if sku_descriptions is None:
        sku_descriptions = {}
    if sku_costs is None:
        sku_costs = {}
    if po_data is None:
        po_data = {}

    soh_data = ref_data["soh_data"]
    bom_index = ref_data["bom_index"]
    stock_take_data = ref_data["stock_take_data"]

    # Build investigation DataFrame (SOH + stock take + PO)
    inv_rows = []
    for sku in skus:
        key = sku.upper()
        soh = soh_data.get(key, {})
        lc = stock_take_data.get(key, {})
        po = po_data.get(key, {})
        avg_cost = sku_costs.get(sku, 0) or 0

        counted_qty = lc.get("Last_Counted_Qty")
        st_variance = lc.get("Last_ST_Variance")
        last_variance_amount = (st_variance * avg_cost) if st_variance is not None else None

        inv_rows.append({
            "Code": sku,
            "Product Description": sku_descriptions.get(sku, ""),
            "AverageCost": avg_cost,
            "SOH Total OnHand": soh.get("SOH_Total_OnHand"),
            "SOH Per Location": soh.get("SOH_Per_Location", "NOT FOUND"),
            "Last StockTake Ref": lc.get("Last_StockTake_Ref", ""),
            "Last StockTake Date": lc.get("Last_StockTake_Date", ""),
            "Last StockTake Location": lc.get("Last_StockTake_Location", ""),
            "Last System SOH": lc.get("Last_System_SOH"),
            "Last Counted Qty": counted_qty,
            "Last ST Variance": st_variance,
            "Last ST Variance Amount": last_variance_amount,
            "Last PO Number": po.get("Last_PO_Number", ""),
            "Last PO Date": po.get("Last_PO_Date", ""),
            "Last PO Qty": po.get("Last_PO_Qty"),
            "Last PO Location": po.get("Last_PO_Location", ""),
        })

    df_investigation = pd.DataFrame(inv_rows)

    # Build BOM DataFrame (exploded: one row per SKU per parent)
    bom_rows = []
    for sku in skus:
        key = sku.upper()
        bom_entries = bom_index.get(key, [])
        for entry in bom_entries:
            bom_rows.append({
                "Code": sku,
                "Product Description": sku_descriptions.get(sku, ""),
                "Parent_SKU": entry["ParentSKU"],
                "Parent_Name": entry["ParentName"],
                "Qty_In_BOM": entry["Quantity"],
                "Qty_To_Produce": entry["QuantityToProduce"],
            })

    df_bom = pd.DataFrame(bom_rows) if bom_rows else pd.DataFrame(
        columns=["Code", "Product Description", "Parent_SKU", "Parent_Name",
                 "Qty_In_BOM", "Qty_To_Produce"]
    )

    return df_investigation, df_bom


# -------------------------
# PDF Helpers
# -------------------------
def detect_vertical_splits_from_lines(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        p = pdf.pages[0]
        W, H = p.width, p.height

        text = p.extract_words(use_text_flow=True) or []
        if text:
            y_tops = [w["top"] for w in text][:40]
            y_header_top = min(y_tops) if y_tops else H * 0.10
        else:
            y_header_top = H * 0.10

        band_top = max(0, y_header_top - 30)
        band_bottom = min(H, band_top + 120)

        verts = []
        for ln in p.lines or []:
            x0, x1, y0, y1 = ln["x0"], ln["x1"], ln["y0"], ln["y1"]
            if abs(x1 - x0) < 0.5:
                y_top_l = min(y0, y1)
                y_bot_l = max(y0, y1)
                if not (y_bot_l < band_top or y_top_l > band_bottom):
                    verts.append(round((x0 + x1) / 2.0, 2))

        verts = sorted(verts)
        dedup = []
        for x in verts:
            if not dedup or abs(x - dedup[-1]) > 2.5:
                dedup.append(x)

        if len(dedup) >= 6:
            return dedup, (0, max(0, band_top - 10), W, H - 6)
        return None, (0, 0, W, H - 6)


def to_num(s):
    return (
        s.astype(str)
        .str.replace("\u00A0", " ", regex=False)
        .str.replace(" ", "", regex=True)
        .str.replace(",", "", regex=True)
        .str.replace(r"[^0-9.\-]", "", regex=True)
        .replace({"": np.nan, ".": np.nan, "-.": np.nan})
        .astype(float)
    )


def extract_and_refine_pdf(pdf_path):
    splits, area_bbox = detect_vertical_splits_from_lines(pdf_path)

    if not splits:
        MANUAL_PERC = [0.03, 0.12, 0.70, 0.78, 0.86, 0.93]
        with pdfplumber.open(pdf_path) as pdf:
            W, H = pdf.pages[0].width, pdf.pages[0].height
        splits = [round(W * p, 2) for p in MANUAL_PERC]
        area_bbox = (0, 0, W, H - 6)

    left, top, right, bottom = area_bbox
    table_area = f"{left:.2f},{top:.2f},{right:.2f},{bottom:.2f}"
    columns_csv = ",".join(f"{x:.2f}" for x in splits)

    tables = camelot.read_pdf(
        pdf_path,
        pages="1-end",
        flavor="stream",
        table_areas=[table_area],
        columns=[columns_csv],
        strip_text=" \n",
        edge_tol=500,
        row_tol=10,
        column_tol=12,
    )

    dfs = [pd.DataFrame(t.df).astype(str) for t in tables]
    max_cols = max((d.shape[1] for d in dfs), default=0)
    aligned = []
    for d in dfs:
        if d.shape[1] < max_cols:
            for i in range(d.shape[1] + 1, max_cols + 1):
                d[f"__pad{i}__"] = ""
        d = d.iloc[:, :max_cols]
        d.columns = [f"Col{i}" for i in range(1, max_cols + 1)]
        aligned.append(d)

    raw = pd.concat(aligned, ignore_index=True) if aligned else pd.DataFrame()

    df_refined = raw.copy()
    df_refined = df_refined.iloc[1:].reset_index(drop=True)
    df_refined = df_refined.drop(df_refined.columns[0], axis=1)

    TARGET = [
        "#", "Code", "Product Description", "Unit",
        "Old Quantity", "Stocktake Quantity", "Variance",
    ]
    if df_refined.shape[1] < 7:
        for i in range(df_refined.shape[1], 7):
            df_refined[f"__pad{i}__"] = np.nan
    df_refined = df_refined.iloc[:, :7].copy()
    df_refined.columns = TARGET

    df_refined = df_refined.map(lambda x: x.strip() if isinstance(x, str) else x)
    df_refined = df_refined.replace({"": np.nan, "nan": np.nan})

    def is_numeric_str(s):
        if not isinstance(s, str):
            return False
        return bool(re.fullmatch(r"\d+", s.strip()))

    mask_hash_ok = df_refined["#"].astype(str).apply(is_numeric_str)
    df_refined = df_refined[mask_hash_ok].copy()
    df_refined = df_refined.dropna(how="all").reset_index(drop=True)

    size_like = re.compile(
        r"^(?:[A-Za-z]*\s*\d+(?:\.\d+)?\s*(?:ml|g|kg|l)|\d+(?:\.\d+)?\s*(?:ml|g|kg|l))$",
        re.IGNORECASE,
    )
    cont_mask = df_refined["Code"].isna() & (
        df_refined["Product Description"]
        .fillna("")
        .str.replace(" ", "", regex=False)
        .str.match(size_like)
    )
    df_refined = df_refined[~cont_mask].reset_index(drop=True)

    for c in ["Old Quantity", "Stocktake Quantity", "Variance"]:
        df_refined[c] = to_num(df_refined[c])

    df_refined["Code"] = df_refined["Code"].astype("string")
    return df_refined


# -------------------------
# Grouping & Merging
# -------------------------
def group_and_sort(df_refined):
    df_grouped = (
        df_refined.groupby(["Code", "Product Description"])
        .agg({
            "Old Quantity": "sum",
            "Stocktake Quantity": "sum",
            "Variance": "sum",
        })
        .reset_index()
    )
    return df_grouped.sort_values(by="Variance", ascending=False).reset_index(drop=True)


def merge_with_inventory(df_grouped_sorted, df_inventory):
    df_inv = df_inventory.copy()
    df_inv["ProductCode"] = df_inv["ProductCode"].astype(str).str.strip()

    if "AverageCost" in df_inv.columns:
        df_inv["AverageCost"] = pd.to_numeric(df_inv["AverageCost"], errors="coerce").fillna(0)
    else:
        df_inv["AverageCost"] = 0

    if "Category" in df_inv.columns:
        df_inv["Category"] = df_inv["Category"].astype(str).fillna("")
    else:
        df_inv["Category"] = ""

    if "DefaultUnitOfMeasure" in df_inv.columns:
        df_inv["Unit"] = df_inv["DefaultUnitOfMeasure"].astype(str).fillna("")
    else:
        df_inv["Unit"] = ""

    if "CommaDelimitedTags" in df_inv.columns:
        df_inv["Tags"] = df_inv["CommaDelimitedTags"].astype(str).fillna("")
    else:
        df_inv["Tags"] = ""

    merge_cols = ["ProductCode", "AverageCost", "Category", "Unit", "Tags"]
    df_merged = pd.merge(
        df_grouped_sorted,
        df_inv[merge_cols],
        left_on="Code",
        right_on="ProductCode",
        how="left",
    )

    df_merged["AverageCost"] = df_merged["AverageCost"].fillna(0)
    df_merged["Category"] = df_merged["Category"].fillna("")
    df_merged["Unit"] = df_merged["Unit"].fillna("")
    df_merged["Tags"] = df_merged["Tags"].fillna("")

    df_merged["Amount of Variance"] = df_merged["AverageCost"] * df_merged["Variance"]
    df_merged["Absolute Amount Variance"] = df_merged["Amount of Variance"].abs()

    df_merged = df_merged.sort_values(by="Absolute Amount Variance", ascending=False).reset_index(drop=True)

    total_abs_variance = df_merged["Absolute Amount Variance"].sum()
    if total_abs_variance > 0:
        df_merged["Cumulative Absolute Variance"] = df_merged["Absolute Amount Variance"].cumsum()
        df_merged["Cumulative Percentage"] = (
            df_merged["Cumulative Absolute Variance"] / total_abs_variance
        ) * 100
        prev_cum = df_merged["Cumulative Percentage"].shift(1, fill_value=0)
        df_merged["Pareto Category"] = prev_cum.apply(
            lambda x: "Top 80% Variance" if x < 80 else "Other"
        )
    else:
        df_merged["Cumulative Absolute Variance"] = 0
        df_merged["Cumulative Percentage"] = 0
        df_merged["Pareto Category"] = "No Variance"

    return df_merged


# -------------------------
# Excel Export
# -------------------------
def autofit_sheet(ws):
    """Auto-fit all column widths in a worksheet."""
    for column in ws.columns:
        col_cells = list(column)
        max_length = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_length + 2, 60)


def build_excel(df_refined, df_merged, df_investigation=None, df_bom=None):
    """Build the Excel workbook in memory and return bytes."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # --- Sheet 1: Grouped & Sorted Data ---
        df_export = df_merged.drop(columns=["ProductCode"], errors="ignore")
        col_order = [
            "Code", "Product Description", "Category", "Unit", "Tags",
            "Old Quantity", "Stocktake Quantity", "Variance",
            "AverageCost", "Amount of Variance", "Absolute Amount Variance",
            "Pareto Category",
        ]
        col_order = [c for c in col_order if c in df_export.columns]
        df_export = df_export[col_order]

        df_export.to_excel(writer, sheet_name="Grouped & Sorted Data", index=False)
        ws_grouped = writer.sheets["Grouped & Sorted Data"]
        ws_grouped.freeze_panes = "D2"

        if "Code" in df_export.columns:
            idx = df_export.columns.get_loc("Code") + 1
            for cell in ws_grouped.iter_rows(min_row=2, min_col=idx, max_col=idx):
                cell[0].number_format = "@"
        if "Amount of Variance" in df_export.columns:
            idx = df_export.columns.get_loc("Amount of Variance") + 1
            for cell in ws_grouped.iter_rows(min_row=2, min_col=idx, max_col=idx):
                cell[0].number_format = "R#,##0.00"
        if "Absolute Amount Variance" in df_export.columns:
            idx = df_export.columns.get_loc("Absolute Amount Variance") + 1
            for cell in ws_grouped.iter_rows(min_row=2, min_col=idx, max_col=idx):
                cell[0].number_format = "R#,##0.00"
        autofit_sheet(ws_grouped)

        # --- Sheet 2: Investigation ---
        if df_investigation is not None and not df_investigation.empty:
            inv_drop_cols = [c for c in ["Product Description", "AverageCost"]
                            if c in df_investigation.columns]
            inv_clean = df_investigation.drop(columns=inv_drop_cols, errors="ignore")
            inv_export = pd.merge(
                df_export[["Code", "Product Description", "Variance",
                           "AverageCost", "Amount of Variance", "Pareto Category"]],
                inv_clean,
                on="Code",
                how="inner",
            )
            inv_export.to_excel(writer, sheet_name="Stock Movements", index=False)
            ws_inv = writer.sheets["Stock Movements"]
            ws_inv.freeze_panes = "C2"
            if "Code" in inv_export.columns:
                idx = inv_export.columns.get_loc("Code") + 1
                for cell in ws_inv.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    cell[0].number_format = "@"
            autofit_sheet(ws_inv)

        # --- Sheet 3: BOM Analysis ---
        if df_bom is not None and not df_bom.empty:
            df_bom.to_excel(writer, sheet_name="BOM Analysis", index=False)
            ws_bom = writer.sheets["BOM Analysis"]
            ws_bom.freeze_panes = "C2"
            if "Code" in df_bom.columns:
                idx = df_bom.columns.get_loc("Code") + 1
                for cell in ws_bom.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    cell[0].number_format = "@"
            autofit_sheet(ws_bom)

        # --- Sheet 4: Refined Data ---
        df_refined.to_excel(writer, sheet_name="Refined Data", index=False)
        ws_refined = writer.sheets["Refined Data"]
        ws_refined.freeze_panes = "D2"
        if "Code" in df_refined.columns:
            idx = df_refined.columns.get_loc("Code") + 1
            for cell in ws_refined.iter_rows(min_row=2, min_col=idx, max_col=idx):
                cell[0].number_format = "@"
        autofit_sheet(ws_refined)

    return output.getvalue()


# =========================================================
# UI
# =========================================================
account_id, api_key = get_api_credentials()

if not account_id or not api_key:
    st.error(
        "Cin7 Core API credentials not found. "
        "Set `CIN7_ACCOUNT_ID` and `CIN7_API_KEY` in your `.env` file (local) "
        "or in Streamlit secrets (cloud deployment)."
    )
    st.stop()

# -------------------------
# Sidebar: Reference Data Loader
# -------------------------
with st.sidebar:
    st.header("Reference Data")

    ref_data = st.session_state.get("ref_data")
    if ref_data:
        st.success(f"Loaded at {ref_data['loaded_at']}")
        st.caption(
            f"SOH: {len(ref_data['soh_data']):,} SKUs | "
            f"BOM: {len(ref_data['bom_index']):,} components | "
            f"ST: {len(ref_data['stock_take_data']):,} SKUs"
        )
        load_label = "Refresh Reference Data"
    else:
        st.warning("Not loaded yet")
        st.caption("Load reference data to enable investigation. This fetches SOH, BOM, and stock take history for all products.")
        load_label = "Load Reference Data"

    load_btn = st.button(load_label, use_container_width=True)

    if load_btn:
        progress_bar = st.progress(0)
        status_text = st.empty()
        client = Cin7Client(account_id, api_key)
        try:
            st.session_state.ref_data = load_reference_data(client, progress_bar, status_text)
            # Clear any previous investigation since ref data changed
            for key in ["df_investigation", "df_bom", "po_data", "po_debug"]:
                st.session_state.pop(key, None)
        except requests.HTTPError as e:
            st.error(f"API error: {e}")
        except requests.ConnectionError:
            st.error("Could not connect to Cin7 Core API.")
        finally:
            progress_bar.empty()
            status_text.empty()
        st.rerun()

    # PO Debug expander in sidebar
    po_debug = st.session_state.get("po_debug")
    if po_debug:
        with st.expander("PO Debug", expanded=False):
            st.write(f"**SKUs queried:** {po_debug.get('skus_queried', '?')}")
            st.write(f"**SKUs with PO data:** {po_debug.get('skus_with_po', '?')}")
            st.write(f"**Total movements seen:** {po_debug.get('total_movements_seen', '?')}")
            if po_debug.get("errors"):
                st.write("**Errors:**")
                for err in po_debug["errors"][:10]:
                    st.write(f"- {err}")


# -------------------------
# Step 1: File Upload & Variance Analysis
# -------------------------
with st.form("inputs"):
    st.subheader("Upload your Stocktake PDF")
    pdf_upload = st.file_uploader("Stocktake Variance Table PDF", type=["pdf"])
    run_btn = st.form_submit_button("Run Analysis")

if run_btn:
    if not pdf_upload:
        st.error("Please upload the Stocktake PDF.")
        st.stop()

    # Clear previous results
    for key in ["df_refined", "df_merged", "df_investigation", "df_bom", "excel_bytes"]:
        st.session_state.pop(key, None)

    with st.spinner("Extracting data from PDF..."):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(pdf_upload.read())
            tmp_pdf_path = tmp.name
        try:
            st.session_state.df_refined = extract_and_refine_pdf(tmp_pdf_path)
        finally:
            os.unlink(tmp_pdf_path)

    with st.spinner("Fetching inventory from Cin7 Core API..."):
        try:
            df_inventory = fetch_inventory_from_api(account_id, api_key)
        except requests.HTTPError as e:
            st.error(f"Cin7 API error: {e}")
            st.stop()
        except requests.ConnectionError:
            st.error("Could not connect to Cin7 Core API. Check your internet connection.")
            st.stop()

    with st.spinner("Grouping and merging..."):
        df_grouped_sorted = group_and_sort(st.session_state.df_refined)
        st.session_state.df_merged = merge_with_inventory(df_grouped_sorted, df_inventory)

    # Build Excel without investigation
    st.session_state.excel_bytes = build_excel(
        st.session_state.df_refined,
        st.session_state.df_merged,
    )

# --- Show results if available ---
if "df_merged" not in st.session_state:
    st.info("Upload the Stocktake PDF above and click **Run Analysis**.")
    st.stop()

df_merged = st.session_state.df_merged
df_refined = st.session_state.df_refined
df_investigation = st.session_state.get("df_investigation")
df_bom = st.session_state.get("df_bom")

st.divider()

# Summary metrics
total_variance_amount = df_merged["Amount of Variance"].sum()
top80_count = len(df_merged[df_merged["Pareto Category"] == "Top 80% Variance"])
total_products = len(df_merged)

c1, c2, c3 = st.columns(3)
c1.metric("Total Products", f"{total_products:,}")
c2.metric("Total Variance Amount", f"R{total_variance_amount:,.2f}")
c3.metric("Top 80% Variance Items", f"{top80_count:,}")

# -------------------------
# Investigation Slider (reactive, no button needed)
# -------------------------
st.divider()
st.subheader("Variance Investigation")

ref_data = st.session_state.get("ref_data")

if not ref_data:
    st.info("Load reference data from the sidebar first to enable investigation.")
else:
    max_skus = len(df_merged)
    default_n = min(25, max_skus)

    top_n = st.slider(
        "Number of top SKUs to investigate (sorted by absolute variance):",
        min_value=5,
        max_value=max_skus,
        value=default_n,
        step=5,
        key="top_n_slider",
    )

    # Build SKU lists for current top N
    top_rows = df_merged.head(top_n)
    top_skus = top_rows["Code"].dropna().astype(str).str.strip().tolist()
    sku_desc_map = dict(zip(
        top_rows["Code"].astype(str).str.strip(),
        top_rows["Product Description"].fillna(""),
    ))
    sku_cost_map = dict(zip(
        top_rows["Code"].astype(str).str.strip(),
        top_rows["AverageCost"].fillna(0),
    ))

    # PO receiving: fetch via product movements (1 call per SKU)
    po_data = st.session_state.get("po_data", {})
    fetched_count = sum(1 for s in top_skus if s.upper() in po_data)

    col_po_btn, col_po_status = st.columns([1, 2])
    with col_po_btn:
        fetch_po_btn = st.button(
            "Fetch PO Data" if not po_data else "Refresh PO Data",
            help=f"Fetch latest PO for each of the top {top_n} SKUs ({top_n} API calls)",
        )
    with col_po_status:
        if po_data:
            st.caption(f"PO data: {fetched_count}/{len(top_skus)} SKUs loaded")
        else:
            st.caption("No PO data loaded yet. Click to fetch.")

    if fetch_po_btn:
        po_progress = st.progress(0)
        po_status = st.empty()
        po_status.text(f"Fetching PO movements for {len(top_skus)} SKUs...")
        client = Cin7Client(account_id, api_key)

        def po_prog_cb(done, total):
            po_progress.progress(
                min(done / max(total, 1), 1.0),
                text=f"Product movements: {done}/{total} SKUs",
            )

        try:
            new_po_data, po_debug = fetch_po_data_via_movements(
                client, top_skus, progress_cb=po_prog_cb,
            )
            # Merge new results into existing cache
            po_data.update(new_po_data)
            st.session_state.po_data = po_data
            st.session_state.po_debug = po_debug
        except requests.HTTPError as e:
            st.warning(f"PO fetch partially failed: {e}")

        po_progress.empty()
        po_status.empty()
        st.rerun()

    # Build investigation instantly from pre-loaded data + PO cache
    df_investigation, df_bom = build_investigation_from_ref(
        ref_data, top_skus, sku_desc_map, sku_cost_map, po_data=po_data,
    )
    st.session_state.df_investigation = df_investigation
    st.session_state.df_bom = df_bom

    # Rebuild Excel with investigation + BOM data
    st.session_state.excel_bytes = build_excel(
        df_refined, df_merged, df_investigation, df_bom,
    )

# Data tables
tab1, tab2, tab3, tab4 = st.tabs([
    "Grouped & Sorted Data", "Stock Movements", "BOM Analysis", "Refined Data",
])

with tab1:
    display_cols = [
        "Code", "Product Description", "Category", "Unit", "Tags",
        "Old Quantity", "Stocktake Quantity", "Variance",
        "AverageCost", "Amount of Variance", "Absolute Amount Variance",
        "Pareto Category",
    ]
    display_cols = [c for c in display_cols if c in df_merged.columns]
    st.dataframe(
        df_merged.drop(columns=["ProductCode"], errors="ignore")[display_cols],
        use_container_width=True,
        hide_index=True,
    )

with tab2:
    if df_investigation is not None and not df_investigation.empty:
        st.dataframe(df_investigation, use_container_width=True, hide_index=True)
    elif not ref_data:
        st.info("Load reference data from the sidebar to enable investigation.")
    else:
        st.info("Adjust the slider above to investigate top SKUs.")

with tab3:
    if df_bom is not None and not df_bom.empty:
        st.dataframe(df_bom, use_container_width=True, hide_index=True)
    elif not ref_data:
        st.info("Load reference data from the sidebar to see BOM relationships.")
    else:
        st.info("No BOM relationships found for the selected SKUs.")

with tab4:
    st.dataframe(df_refined, use_container_width=True, hide_index=True)

# --- Download ---
st.divider()
timestamp = datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
inv_label = ""
if df_investigation is not None and not df_investigation.empty:
    inv_label = f" (incl. {len(df_investigation)} investigated)"
st.download_button(
    label=f"Download Variance Analysis XLSX{inv_label}",
    data=st.session_state.excel_bytes,
    file_name=f"stocktake_variance_analysis_{timestamp}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
