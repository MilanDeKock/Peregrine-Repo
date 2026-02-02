import io
import re
from datetime import datetime
from pytz import timezone

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border

# -------------------------
# Configuration & Styling
# -------------------------
TZ = timezone("Africa/Johannesburg")
PEREGRINE_BLUE = "00A9E0"   # RGB(0,169,224)
DIGISCALE_GREEN = "3E5D3E"  # RGB(62,93,62)
README_RED = "FF0000"       # RGB(255,0,0)

st.set_page_config(page_title="Systems Reconciliation | Cin7 Report Summary", page_icon="🧾", layout="centered")

st.markdown(
    """
    <style>
    h1, .stMarkdown h1, .stCaption p { white-space: nowrap !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

# -------------------------
# Helpers
# -------------------------
def style_and_autofit_sheet(ws, rgb_hex="00A9E0"):
    rgb = rgb_hex.replace("#", "").upper()
    argb = ("00" + rgb) if len(rgb) == 6 else (rgb if len(rgb) == 8 else "00A9E0")
    header_fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    no_border = Border()

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = no_border

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)

def create_readme(writer):
    wb = writer.book
    ws = wb.create_sheet("README", 0)
    ws.sheet_properties.tabColor = README_RED
    content = [
        ["SHEET NAME", "PURPOSE / ACTION REQUIRED"],
        ["Yoco Products - Not in Core", "Items found in Yoco but missing from Core. Action: Add to Core or fix SKU."],
        ["Yoco Products - Price Mismatch", "Selling prices don't match. Action: Sync prices between systems."],
        ["Core Sellable - Not in Yoco", "Items marked 'Sellable' in Core but missing from Yoco. Action: Add to Yoco."],
        ["Modifiers - Not in Core", "Modifier items missing from Core. Action: Ensure all modifiers have Core PLUs."],
        ["Digiscale - Not in Core", "Items on the Digiscale list not found in Core. Action: Update Core inventory."],
        ["Digiscale - Price Mismatch", "Scale prices vs Core prices. Action: Update scale pricing."]
    ]
    for row in content:
        ws.append(row)
    style_and_autofit_sheet(ws, rgb_hex=README_RED)

def read_core_csv(uploaded):
    try:
        return pd.read_csv(uploaded, encoding="utf-8-sig")
    except UnicodeDecodeError:
        return pd.read_csv(uploaded)

def read_excel_sheet(uploaded_file, sheet_name):
    xls = pd.ExcelFile(uploaded_file)
    if sheet_name not in xls.sheet_names:
        raise KeyError(f"Sheet '{sheet_name}' not found. Found: {xls.sheet_names}")
    return xls.parse(sheet_name)

def norm(s):
    return re.sub(r'[^a-z0-9]+', ' ', str(s).lower()).strip()

def map_first_available(df, candidates, require_tokens=None):
    norm_cols = {norm(c): c for c in df.columns}
    for k in candidates:
        if k in norm_cols:
            return norm_cols[k]
    if require_tokens:
        toks = require_tokens if isinstance(require_tokens, (list, tuple)) else [require_tokens]
        for nc, raw in norm_cols.items():
            words = nc.split()
            if all(t in words for t in toks):
                return raw
    return None

def norm_plu_series(s):
    return s.astype(str).str.strip().str.upper()

# -------------------------
# UI: File Upload
# -------------------------
st.title("🧾 Systems Reconciliation | Cin7 Report Summary")
st.caption("BULK sheet + Modifiers + Digiscale checks → one XLSX download")

with st.form("inputs"):
    st.subheader("Please upload reconciliation files")
    uploaded_files = st.file_uploader(
        "Upload Core CSV, Yoco XLSX, Modifiers, and Digiscale CSV all at once:",
        accept_multiple_files=True,
        type=["csv", "xlsx", "xls"]
    )
    run_btn = st.form_submit_button("Run reconciliation")

if not run_btn:
    st.info("Upload required files above and click **Run reconciliation**.")
    st.stop()

# -------------------------
# Detection
# -------------------------
core_file = yoco_file = modifiers_alt = digiscale_file = None
for f in uploaded_files or []:
    name = f.name.lower()
    if "inventorylist" in name: core_file = f
    elif "peregrine" in name and "farm" in name and "bulk" in name: yoco_file = f
    elif "modifier" in name: modifiers_alt = f
    elif "digiscale" in name or "plu listing" in name: digiscale_file = f

if not core_file or not yoco_file:
    st.error("You must include both a Core Inventory CSV and a Yoco BULK XLSX.")
    st.stop()

# -------------------------
# Data Loading & Recon
# -------------------------
core_df = read_core_csv(core_file)
yoco_df = read_excel_sheet(yoco_file, "Bulk Site Values")

yoco_df["PLU_norm"] = norm_plu_series(yoco_df["Product PLU"])
core_df["ProductCode_norm"] = norm_plu_series(core_df["ProductCode"])

# 1) Yoco vs Core
prod_merge = yoco_df.merge(core_df[["ProductCode", "PriceTier1", "ProductCode_norm"]], left_on="PLU_norm", right_on="ProductCode_norm", how="left")
yoco_not_in_core = prod_merge[prod_merge["ProductCode"].isna()][["Product PLU", "Name & Variants"]]
y_price = pd.to_numeric(prod_merge["Selling Price"], errors="coerce")
c_price = pd.to_numeric(prod_merge["PriceTier1"], errors="coerce")
yoco_price_mismatch = prod_merge.loc[prod_merge["ProductCode"].notna() & (y_price.ne(c_price)), ["Product PLU", "Name & Variants", "Selling Price", "PriceTier1"]].rename(columns={"Selling Price": "Yoco Price", "PriceTier1": "Core Price"})

# 2) Core Sellable vs Yoco
sellable_mask = core_df["Sellable"].astype(str).str.strip().str.lower().eq("yes")
core_sellable_merge = core_df.loc[sellable_mask].merge(yoco_df[["Product PLU", "PLU_norm"]], left_on="ProductCode_norm", right_on="PLU_norm", how="left")
core_sellable_not_in_yoco = core_sellable_merge[core_sellable_merge["Product PLU"].isna()][["ProductCode", "Name"]].rename(columns={"ProductCode": "Core Product Code", "Name": "Core Name"})

# 3) Modifiers (Original Logic Reinstated)
mod_df = None
try:
    xls = pd.ExcelFile(yoco_file)
    if "Modifier Items - Template" in xls.sheet_names:
        mod_df = xls.parse("Modifier Items - Template")
except: pass

if mod_df is None and modifiers_alt is not None:
    try:
        xls_alt = pd.ExcelFile(modifiers_alt)
        if "Modifier Items - Template" in xls_alt.sheet_names:
            mod_df = xls_alt.parse("Modifier Items - Template")
        else:
            cand = next((s for s in xls_alt.sheet_names if "modifier" in s.lower()), xls_alt.sheet_names[0])
            mod_df = xls_alt.parse(cand)
    except Exception as e:
        st.error(f"Failed to read Modifiers: {e}"); st.stop()

if mod_df is None:
    st.error("Could not locate 'Modifier Items - Template'."); st.stop()

mod_code_col = map_first_available(mod_df, candidates=["modifier items plu product modifier", "modifier items plu", "modifier plu", "product modifier plu", "modifier product plu", "plu"], require_tokens=["modifier", "plu"]) or map_first_available(mod_df, candidates=["plu"])
mod_name_col = map_first_available(mod_df, candidates=["modifier name", "name"], require_tokens=["modifier", "name"]) or map_first_available(mod_df, candidates=["name"])
mod_type_col = map_first_available(mod_df, candidates=["type products options", "type products options ", "type"])
mod_item_col = map_first_available(mod_df, candidates=["modifier item", "modifier items", "item"], require_tokens=["modifier", "item"]) or map_first_available(mod_df, candidates=["item"])

_mod = mod_df.copy()
_mod["_PLU_norm"] = norm_plu_series(_mod[mod_code_col])
mods_merged = _mod.merge(core_df[["ProductCode", "ProductCode_norm"]], left_on="_PLU_norm", right_on="ProductCode_norm", how="left")
mods_not_in_core = mods_merged[mods_merged["ProductCode"].isna()][[mod_name_col, mod_type_col, mod_item_col]].copy()
mods_not_in_core.columns = ["Modifier Name", "Type (Products / Options)", "Modifier Item"]

# 4) Digiscale
digi_not_in_core = pd.DataFrame(columns=["Digiscale SKU", "Digiscale Name"])
digi_price_mismatch = pd.DataFrame(columns=["Digiscale SKU", "Digiscale Name", "Yoco Price", "Core Price"])
if digiscale_file:
    digi_df = read_core_csv(digiscale_file)
    digi_df["_PLU_norm"] = norm_plu_series(digi_df["PLU #"])
    digi_merged = digi_df.merge(core_df[["ProductCode", "PriceTier1", "ProductCode_norm"]], left_on="_PLU_norm", right_on="ProductCode_norm", how="left")
    digi_not_in_core = digi_merged[digi_merged["ProductCode"].isna()][["PLU #", "Description Line 1"]].rename(columns={"PLU #": "Digiscale SKU", "Description Line 1": "Digiscale Name"})
    d_p = pd.to_numeric(digi_merged["Price"], errors="coerce"); c_p = pd.to_numeric(digi_merged["PriceTier1"], errors="coerce")
    digi_price_mismatch = digi_merged.loc[digi_merged["ProductCode"].notna() & (d_p.ne(c_p)), ["PLU #", "Description Line 1", "Price", "PriceTier1"]].rename(columns={"PLU #": "Digiscale SKU", "Description Line 1": "Digiscale Name", "Price": "Yoco Price", "PriceTier1": "Core Price"})

# -------------------------
# Export Phase 1
# -------------------------
out_xlsx = io.BytesIO()
with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
    create_readme(writer)
    yoco_not_in_core.to_excel(writer, index=False, sheet_name="Yoco Products - Not in Core")
    yoco_price_mismatch.to_excel(writer, index=False, sheet_name="Yoco Products - Price Mismatch")
    core_sellable_not_in_yoco.to_excel(writer, index=False, sheet_name="Core Sellable - Not in Yoco")
    mods_not_in_core.to_excel(writer, index=False, sheet_name="Modifiers - Not in Core")
    digi_not_in_core.to_excel(writer, index=False, sheet_name="Digiscale - Not in Core")
    digi_price_mismatch.to_excel(writer, index=False, sheet_name="Digiscale - Price Mismatch")
    wb = writer.book
    for s in wb.sheetnames:
        if s == "README": continue
        color = DIGISCALE_GREEN if "Digiscale" in s else PEREGRINE_BLUE
        style_and_autofit_sheet(wb[s], rgb_hex=color)

st.success("✅ Phase 1 Complete.")
st.download_button("⬇️ Download Recon XLSX", data=out_xlsx.getvalue(), file_name=f"core_yoco_digiscale_recon_{datetime.now(TZ).strftime('%Y%m%d')}.xlsx")

with st.expander("Show recon summary"):
    st.write({
        "Yoco Products - Not in Core": len(yoco_not_in_core),
        "Yoco Products - Price Mismatch": len(yoco_price_mismatch),
        "Core Sellable - Not in Yoco": len(core_sellable_not_in_yoco),
        "Modifiers - Not in Core": len(mods_not_in_core),
        "Digiscale - Not in Core": len(digi_not_in_core),
        "Digiscale - Price Mismatch": len(digi_price_mismatch),
    })

# -------------------------
# Phase 2: Monthly Health Check
# -------------------------
st.divider()
st.subheader("🔍 Phase 2: Monthly Health Check")
health_file = st.file_uploader("Upload Cin7 Health Check XLSX:", type=["xlsx"])

if health_file:
    try:
        df_dupes = pd.read_excel(health_file, sheet_name="Duplicate Sales")
        df_margin = pd.read_excel(health_file, sheet_name="Assembly BOM vs Margin Check")
        df_stock = pd.read_excel(health_file, sheet_name="Stock Adjustment Analysis")
        df_bom = pd.read_excel(health_file, sheet_name="Deleted BOM Lines Check")

        # --- 1. Duplicate Sales ---
        dupe_count = len(df_dupes)

        # --- 2. Margin Analysis (Handling 45.28% vs 0.4528) ---
        # Convert to numeric, errors to NaN
        m_vals = pd.to_numeric(df_margin["GP Margin"], errors='coerce')
        
        # If the max value is <= 1.0, it's likely decimal (0.45), so multiply by 100
        # If it's already > 1.0, we treat it as literal percentage (45.28)
        if m_vals.max() <= 1.0:
            m_vals = m_vals * 100
        
        df_margin["GP_Clean"] = m_vals
        
        neg_margin_count = len(df_margin[df_margin["GP_Clean"] < 0])
        high_margin_count = len(df_margin[df_margin["GP_Clean"] > 80])
        low_margin_df = df_margin.sort_values("GP_Clean").head(5)

        # --- 3. Stock Adjustments ---
        stock_issues = df_stock[df_stock["Unit Cost"].isin([0, 1])]

        # --- 4. Production Integrity ---
        bom_issues = df_bom[df_bom["Status"].astype(str).str.contains("❌ MISSING", na=False)]

        # --- Metrics Display ---
        c1, c2, c3 = st.columns(3)
        c1.metric("Duplicate Sales", f"{dupe_count}")
        c2.metric("R0/R1 Stock Adj", f"{len(stock_issues)}")
        c3.metric("BOM Missing Lines", f"{len(bom_issues)}")

        st.write("### 📊 Margin Breakdown")
        m1, m2 = st.columns(2)
        m1.info(f"🚩 **{neg_margin_count}** items have a Negative Margin.")
        m2.success(f"💎 **{high_margin_count}** items have a Margin > 80%.")

        st.write("**Bottom 5 Margins**")
        st.dataframe(low_margin_df[["ProductSKU", "ProductName", "GP_Clean"]].rename(columns={"GP_Clean": "GP Margin %"}), hide_index=True)

        # --- Updated Email Summary ---
        st.subheader("✉️ Email Summary")
        
        # Calculate summary strings for the email
        margin_summary = f"{neg_margin_count} items identified with negative margins and {high_margin_count} items with margins above 80%."
        
        email_body = f"""Hi Team,

Please find the reconciliation and health check summary for the period:

Yoco|Core|Digiscale Reconciliation Summary:
- Yoco Items not in Core: {len(yoco_not_in_core)}
- Price Mismatches (Yoco vs Core): {len(yoco_price_mismatch)}
- Missing from Yoco (Sellable in Core): {len(core_sellable_not_in_yoco)}
- Modifiers not in Core: {len(mods_not_in_core)}
- Digiscale Items not in Core: {len(digi_not_in_core)}
- Digiscale Price Mismatches: {len(digi_price_mismatch)}

Cin7 Report Summary:
- Duplicate Sales: {dupe_count} rows found.
- Stock Adjustments: {len(stock_issues)} items adjusted at R0 or R1.
- Production Integrity: {len(bom_issues)} assemblies found with missing BOM lines.
- Margin Analysis: {margin_summary}

Please refer to the attached reports for the full details.

Best regards,"""

        st.text_area("Copy and paste this summary into your email body:", value=email_body, height=450)
        
    except Exception as e:
        st.error(f"Error in Phase 2: {e}")
